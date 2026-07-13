"""
Batch OBDH Pure (No Embedding) Calculation
==========================================
Đọc từng chất trong input.json, chạy UHF + OBDH_CL (pure, không embedding),
quét giá trị alpha_c từ 0.01 đến 1.00 (step 0.01).
Ghi kết quả raw và bảng thống kê sai số ra file XLSX bằng công thức Excel.

Usage:
    python run_obdh_pure.py
    python run_obdh_pure.py --input input.json --basis def2-TZVP --output results.xlsx
"""

import argparse
import json
import time
import traceback
import math

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from concurrent.futures import ProcessPoolExecutor
from pyscf import gto, scf
#from pycmf.OBDH import OBMP2_CL as OBDH_CL
from pycmf.OBDH import OBDH_CL
from pycmf.OBDH.stability import stabilize_scf

# ─────────────────────────── helpers ───────────────────────────

def build_mol(name: str, data: dict, basis: str) -> gto.Mole:
    atom_str = "\n".join(
        f"  {sym}  {x:.6f}  {y:.6f}  {z:.6f}"
        for sym, (x, y, z) in data["geometry"]
    )
    mol = gto.Mole()
    mol.atom    = atom_str
    mol.charge  = data.get("charge", 0)
    mol.spin    = data.get("spin", 0)
    mol.basis   = {'default': 'aug-cc-pcvqz', 'H': 'aug-cc-pvqz'}
    mol.verbose = 0
    mol.build()
    return mol


def run_one(name: str, data: dict, basis: str, obdh_kwargs: dict) -> dict:
    result = {
        "molecule":         name,
        "charge":           data.get("charge", 0),
        "spin":             data.get("spin", 0),
        "multiplicity":     data.get("multiplicity", 1),
        "basis":            basis,
        "uhf_energy":       None,
        "obdh_energy":      None,
        "obdh_corr_energy": None,
        "dip_mom":          None,
        "runtime_uhf_s":    None,
        "runtime_obdh_s":   None,
        "converged_uhf":    None,
        "converged_obdh":   None,
        "error":            None,
    }

    try:
        mol = build_mol(name, data, basis)

        # ── UHF ──────────────────────────────────────────────────────────
        t0 = time.time()
        mf = scf.UHF(mol).density_fit()
        mf.run()
        mf = stabilize_scf(mf, max_macro_cycles=10, verbose=False)
        result["runtime_uhf_s"] = round(time.time() - t0, 4)
        result["uhf_energy"]    = mf.e_tot
        result["converged_uhf"] = bool(mf.converged)
        if not mf.converged:
            print(f"    WARNING: UHF không hội tụ cho {name}")
        

        # ── OBDH pure ────────────────────────────────────────────────────
        t1 = time.time()
        calc = OBDH_CL(mf)
        calc.alphaa          = obdh_kwargs.get("alphaa",          (0.53, 0.39))
        calc.thresh          = obdh_kwargs.get("thresh",          1e-6)
        calc.second_order    = obdh_kwargs.get("second_order",    True)
        calc.mom_select      = obdh_kwargs.get("mom_select",      False)
        calc.mom_start_cycle = obdh_kwargs.get("mom_start_cycle", 0)
        calc.use_embed       = False
        calc.use_cl          = False
        calc.run()
        result["runtime_obdh_s"] = round(time.time() - t1, 4)
        result["converged_obdh"] = getattr(calc, "converged", None)

        e_tot = getattr(calc, "ene_tot", None)
        result["obdh_energy"]      = e_tot
        result["obdh_corr_energy"] = (e_tot - mf.e_tot) if e_tot is not None else None
        
        if hasattr(calc, "dip_mom") and calc.dip_mom is not None:
            result["dip_mom"] = round(calc.dip_mom, 6)

    except Exception:
        result["error"] = traceback.format_exc()

    return result


def print_summary(result: dict):
    ok = result.get("error") is None
    print(
        f"[{'OK' if ok else 'FAILED'}] {result.get('molecule', '')[:20]:20s}"
        f"  alpha_c={result.get('alpha_c', 'N/A')}"
        f"  UHF={result.get('uhf_energy')}"
        f"  OBDH={result.get('obdh_energy')}"
        f"  dip={result.get('dip_mom')} D"
        f"  t={result.get('runtime_obdh_s')}s"
    )
    if not ok and result.get("error"):
        for ln in result["error"].strip().splitlines()[-3:]:
            print("    " + ln)


# ──────────────────────────── xlsx ─────────────────────────────

_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_DATA_FONT   = Font(name="Arial", size=10)
_ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
_BORDER      = Border(**{s: Side(style="thin", color="B0B0B0")
                         for s in ("left", "right", "top", "bottom")})
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left",   vertical="center")

def save_xlsx(results: list, path: str):
    if not results:
        return
    wb = openpyxl.Workbook()
    
    # ── SHEET 1: RAW RESULTS ─────────────────────────────────────────
    ws_raw = wb.active
    ws_raw.title = "Raw Results"
    ws_raw.row_dimensions[1].height = 22

    HEADERS = ["Alpha_c", "Molecule", "Charge", "Spin", "Multiplicity", "Basis",
               "UHF Energy (Eh)", "OBDH Energy (Eh)", "Corr Energy (Eh)",
               "Calc Dipole (D)", "Ref Dipole (CCSD(T))", "Error", "Abs Error", "Sq Error",
               "t_UHF (s)", "t_OBDH (s)", "Conv UHF", "Conv OBDH", "Traceback"]
               
    KEYS    = ["alpha_c", "molecule", "charge", "spin", "multiplicity", "basis",
               "uhf_energy", "obdh_energy", "obdh_corr_energy",
               "dip_mom", None, None, None, None, 
               "runtime_uhf_s", "runtime_obdh_s", "converged_uhf", "converged_obdh", "error"]
               
    NUM_FMT = {7: "0.00000000", 8: "0.00000000", 9: "0.00000000", 
               10: "0.000000", 11: "0.000000", 12: "0.000000", 13: "0.000000", 14: "0.000000"}
               
    COL_W   = [10, 22, 8, 6, 13, 14, 18, 18, 18, 18, 20, 15, 15, 15, 11, 11, 11, 12, 40]

    for col, hdr in enumerate(HEADERS, 1):
        c = ws_raw.cell(row=1, column=col, value=hdr)
        c.font, c.fill, c.border, c.alignment = _HEADER_FONT, _HEADER_FILL, _BORDER, _CENTER

    alpha_ranges = {}
    current_alpha = None

    for row_i, r in enumerate(results, 2):
        ac = r["alpha_c"]
        
        # Track start/end rows cho từng alpha_c để phục vụ làm công thức bên sheet thống kê
        if ac != current_alpha:
            if current_alpha is not None:
                alpha_ranges[current_alpha]["end"] = row_i - 1
            current_alpha = ac
            alpha_ranges[current_alpha] = {"start": row_i, "end": row_i}
        else:
            alpha_ranges[current_alpha]["end"] = row_i

        alt = (row_i % 2 == 0)
        for col, key in enumerate(KEYS, 1):
            if col == 11:    # Ref Dipole (CCSD(T)) gán = 0
                val = 0
            elif col == 12:  # Error = Calc - Ref
                val = f"=J{row_i}-K{row_i}"
            elif col == 13:  # Abs Error = ABS(Error)
                val = f"=ABS(L{row_i})"
            elif col == 14:  # Sq Error = Error^2
                val = f"=L{row_i}^2"
            else:
                val = r.get(key)
                
            c = ws_raw.cell(row=row_i, column=col, value=val)
            c.font, c.border, c.alignment = _DATA_FONT, _BORDER, _LEFT
            if alt:
                c.fill = _ALT_FILL
            if col in NUM_FMT:
                c.number_format = NUM_FMT[col]

    for i, w in enumerate(COL_W, 1):
        ws_raw.column_dimensions[get_column_letter(i)].width = w
    ws_raw.freeze_panes = "B2"


    # ── SHEET 2: STATISTICS ──────────────────────────────────────────
    ws_stat = wb.create_sheet(title="Statistics")
    ws_stat.row_dimensions[1].height = 22
    
    stat_headers = ["Alpha_c", "N_Molecules", "Mean Error (ME)", "Max Error (MAX)", "RMSE"]
    stat_col_w = [12, 14, 18, 18, 18]
    
    for col, hdr in enumerate(stat_headers, 1):
        c = ws_stat.cell(row=1, column=col, value=hdr)
        c.font, c.fill, c.border, c.alignment = _HEADER_FONT, _HEADER_FILL, _BORDER, _CENTER
        ws_stat.column_dimensions[get_column_letter(col)].width = stat_col_w[col-1]

    for row_i, (ac, rng) in enumerate(alpha_ranges.items(), 2):
        s = rng["start"]
        e = rng["end"]
        n_mol = e - s + 1
        alt = (row_i % 2 == 0)
        
        # Điền dữ liệu cơ bản
        ws_stat.cell(row=row_i, column=1, value=ac)
        ws_stat.cell(row=row_i, column=2, value=n_mol)
        
        # Link công thức Excel từ Raw Results
        ws_stat.cell(row=row_i, column=3, value=f"=AVERAGE('Raw Results'!L{s}:L{e})")
        ws_stat.cell(row=row_i, column=4, value=f"=MAX('Raw Results'!M{s}:M{e})")
        ws_stat.cell(row=row_i, column=5, value=f"=SQRT(AVERAGE('Raw Results'!N{s}:N{e}))")
        
        for col in range(1, 6):
            c = ws_stat.cell(row=row_i, column=col)
            c.font, c.border, c.alignment = _DATA_FONT, _BORDER, _LEFT
            if alt:
                c.fill = _ALT_FILL
            if col >= 3:
                c.number_format = "0.000000"

    wb.save(path)
    print(f"\nKết quả đã lưu vào: {path} (Sheet thống kê đã được liên kết công thức tự động)")


def save_log(results: list, path: str):
    with open(path, "w") as f:
        for r in results:
            f.write("=" * 70 + "\n")
            for k, v in r.items():
                f.write(f"  {k}: {v}\n")
    print(f"Log chi tiết đã lưu vào: {path}")


# ─────────────────────────── main ──────────────────────────────

# ĐƯA HÀM NÀY RA NGOÀI HÀM MAIN (ĐỨNG ĐỘC LẬP Ở ĐÂY)
def worker_func(task):
    name, data, basis, obdh_kwargs, alpha_c = task
    res = run_one(name, data, basis, obdh_kwargs)
    res["alpha_c"] = alpha_c
    return res


def main():
    t_start = time.time()
    
    parser = argparse.ArgumentParser(description="Batch OBDH Pure Calculation with alpha_c Scan")
    parser.add_argument("--input",     default="/home/giahuy/Code/job/OBDH/geometry/sp_inputs.json")
    parser.add_argument("--basis",     default="aug-cc-pcvqz")
    parser.add_argument("--output",    default="/data/giahuy/Result/OBDH/dipole_moments/sp_results.xlsx")
    parser.add_argument("--log",       default="/data/giahuy/Result/OBDH/dipole_moments/sp_results.log")
    parser.add_argument("--molecules", nargs="*", default=None)
    args = parser.parse_args()

    with open(args.input) as f:
        molecules = json.load(f)

    if args.molecules:
        molecules = {k: v for k, v in molecules.items() if k in args.molecules}

    alpha_c_list = np.round(np.arange(0.01, 1.01, 0.01), 2)
    
    n_mols = len(molecules)
    n_alphas = len(alpha_c_list)
    n_total = n_mols * n_alphas
    
    print(f"Sẽ tính {n_mols} phân tử x {n_alphas} mốc alpha_c = {n_total} jobs | basis={args.basis} | OBDH pure\n" + "-" * 70)

    tasks = []
    for alpha_c in alpha_c_list:
        for name, data in molecules.items():
            obdh_kwargs = {
                "alphaa": (0.53, float(alpha_c)),
                "thresh": 1e-8,
                "second_order": True,
                "mom_select": False,
                "mom_start_cycle": 0,
            }
            tasks.append((name, data, args.basis, obdh_kwargs, float(alpha_c)))

    # Đã cấu hình tối ưu 15 workers (15 phân tử x 4 cores/phân tử = 60 cores)
    num_workers = 10 
    
    results = []
    print(f"Chạy song song {num_workers} tiến trình...")

    # BÊN TRONG HÀM MAIN BÂY GIỜ CHỈ CÒN ĐOẠN GỌI EXECUTOR NÀY
    with ProcessPoolExecutor(max_workers=num_workers) as executor:
        for r in executor.map(worker_func, tasks):
            results.append(r)
            print_summary(r)

    wall   = round(time.time() - t_start, 2)
    n_ok   = sum(1 for r in results if r["error"] is None)
    n_fail = n_total - n_ok
    print("-" * 70)
    print(f"\nHoàn thành: {n_ok}/{n_total} thành công, {n_fail} lỗi, tổng {wall}s\n")

    save_xlsx(results, args.output)
    save_log(results, args.log)


if __name__ == "__main__":
    main()