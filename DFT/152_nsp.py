"""
Batch DFT Dipole Moment Calculation (PBE0 & B3LYP)
====================================================
Đọc từng chất trong input.json (cùng format với nsp_inputs.json), chạy DFT
(RKS cho closed-shell / UKS cho open-shell) với hai functional PBE0 và B3LYP,
tính dipole moment rồi ghi ra file XLSX 3 cột: Molecule, PBE0, B3LYP.

Usage:
    python run_dft_dipole.py
    python run_dft_dipole.py --input nsp_inputs.json --basis def2-TZVPPD --output dipole_results.xlsx
    python run_dft_dipole.py --molecules AlF BF SO2
"""

import argparse
import json
import time
import traceback

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pyscf import gto, dft
from pycmf.OBDH.stability import stabilize_scf

# ─────────────────────────── helpers ───────────────────────────

def build_mol(name: str, data: dict, basis: str) -> gto.Mole:
    atom_str = "\n".join(
        f"  {sym}  {x:.6f}  {y:.6f}  {z:.6f}"
        for sym, (x, y, z) in data["geometry"]
    )
    mol = gto.Mole()
    mol.atom    = atom_str
    mol.charge  = data.get("charge", 0)
    mol.spin    = data.get("spin", 0)
    mol.basis   = basis
    mol.verbose = 0
    mol.build()
    return mol


def make_ks(mol: gto.Mole, xc: str):
    """RKS cho closed-shell (spin=0), UKS cho open-shell."""
    mf = dft.UKS(mol)
    mf = mf.density_fit()
    mf.xc = xc
    return mf


def run_dft(mol: gto.Mole, xc: str):
    mf = make_ks(mol, xc)
    print(f"   => Running SCF {xc}...")
    mf.kernel()
    mf = stabilize_scf(mf, max_macro_cycles=10, verbose=True)
    dm  = mf.make_rdm1()
    dip = mf.dip_moment(mol, dm, unit="Debye", verbose=0)
    dip_norm = float(np.linalg.norm(dip))
    return mf.e_tot, bool(mf.converged), dip_norm


def run_one(name: str, data: dict, basis: str) -> dict:
    result = {
        "molecule":        name,
        "charge":          data.get("charge", 0),
        "spin":            data.get("spin", 0),
        "multiplicity":    data.get("multiplicity", 1),
        "basis":           basis,
        "pbe0_energy":     None,
        "pbe0_dipole":     None,
        "converged_pbe0":  None,
        "runtime_pbe0_s":  None,
        "b3lyp_energy":    None,
        "b3lyp_dipole":    None,
        "converged_b3lyp": None,
        "runtime_b3lyp_s": None,
        "error":           None,
    }

    try:
        mol = build_mol(name, data, basis)

        # ── PBE0 ─────────────────────────────────────────────────────────
        t0 = time.time()
        e_pbe0, conv_pbe0, dip_pbe0 = run_dft(mol, "PBE0")
        result["runtime_pbe0_s"] = round(time.time() - t0, 4)
        result["pbe0_energy"]    = e_pbe0
        result["converged_pbe0"] = conv_pbe0
        result["pbe0_dipole"]    = round(dip_pbe0, 6)
        if not conv_pbe0:
            print(f"    WARNING: PBE0 không hội tụ cho {name}")

        # ── B3LYP ────────────────────────────────────────────────────────
        t1 = time.time()
        e_b3lyp, conv_b3lyp, dip_b3lyp = run_dft(mol, "B3LYP")
        result["runtime_b3lyp_s"] = round(time.time() - t1, 4)
        result["b3lyp_energy"]    = e_b3lyp
        result["converged_b3lyp"] = conv_b3lyp
        result["b3lyp_dipole"]    = round(dip_b3lyp, 6)
        if not conv_b3lyp:
            print(f"    WARNING: B3LYP không hội tụ cho {name}")

    except Exception:
        result["error"] = traceback.format_exc()

    return result


def print_summary(result: dict):
    ok = result["error"] is None
    print(
        f"[{'OK' if ok else 'FAILED'}] {result['molecule']:20s}"
        f"  PBE0={result['pbe0_dipole']} D"
        f"  B3LYP={result['b3lyp_dipole']} D"
    )
    if not ok:
        for ln in result["error"].strip().splitlines()[-3:]:
            print("    " + ln)


# ──────────────────────────── xlsx ─────────────────────────────
# File kết quả chính: 3 cột Molecule / PBE0 / B3LYP theo yêu cầu

_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_DATA_FONT   = Font(name="Arial", size=10)
_ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
_BORDER      = Border(**{s: Side(style="thin", color="B0B0B0")
                         for s in ("left", "right", "top", "bottom")})
_CENTER      = Alignment(horizontal="center", vertical="center")

HEADERS = ["Molecule", "PBE0 (D)", "B3LYP (D)"]
KEYS    = ["molecule", "pbe0_dipole", "b3lyp_dipole"]
NUM_FMT = {2: "0.000000", 3: "0.000000"}
COL_W   = [22, 14, 14]


def save_xlsx(results: list, path: str):
    if not results:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dipole Moments"
    ws.row_dimensions[1].height = 22

    for col, hdr in enumerate(HEADERS, 1):
        c = ws.cell(1, col, hdr)
        c.font, c.fill, c.border, c.alignment = _HEADER_FONT, _HEADER_FILL, _BORDER, _CENTER

    for row_i, r in enumerate(results, 2):
        alt = (row_i % 2 == 0)
        for col, key in enumerate(KEYS, 1):
            c = ws.cell(row_i, col, r.get(key))
            c.font      = _DATA_FONT
            c.border    = _BORDER
            c.alignment = _CENTER
            if alt:
                c.fill = _ALT_FILL
            if col in NUM_FMT:
                c.number_format = NUM_FMT[col]

    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"\nKết quả đã lưu vào: {path}")


def save_log(results: list, path: str):
    with open(path, "w") as f:
        for r in results:
            f.write("=" * 70 + "\n")
            for k, v in r.items():
                f.write(f"  {k}: {v}\n")
    print(f"Log chi tiết đã lưu vào: {path}")


# ─────────────────────────── main ──────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Batch DFT Dipole Moment (PBE0 & B3LYP)")
    parser.add_argument("--input",     default="/home/giahuy/Code/job/OBDH/geometry/sp_inputs.json")
    parser.add_argument("--basis",     default="aug-pc-3")
    parser.add_argument("--output",    default="/data/giahuy/Result/DFT/dipole_moments/sp_results.xlsx")
    parser.add_argument("--log",       default="/data/giahuy/Result/DFT/dipole_moments/results.log")
    parser.add_argument("--molecules", nargs="*", default=None)
    args = parser.parse_args()

    with open(args.input) as f:
        molecules = json.load(f)

    if args.molecules:
        molecules = {k: v for k, v in molecules.items() if k in args.molecules}

    n_total = len(molecules)
    print(f"Sẽ tính {n_total} phân tử  |  basis={args.basis}  |  DFT: PBE0 & B3LYP\n" + "-" * 70)

    results = []
    t_start = time.time()

    for idx, (name, data) in enumerate(molecules.items(), 1):
        print(f"[{idx}/{n_total}] {name} ...", flush=True)
        result = run_one(name, data, args.basis)
        results.append(result)
        print_summary(result)

    wall   = round(time.time() - t_start, 2)
    n_ok   = sum(1 for r in results if r["error"] is None)
    n_fail = n_total - n_ok
    print("-" * 70)
    print(f"\nHoàn thành: {n_ok}/{n_total} thành công, {n_fail} lỗi, tổng {wall}s\n")

    save_xlsx(results, args.output)
    save_log(results, args.log)


if __name__ == "__main__":
    main()   